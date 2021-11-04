#!/usr/bin/env python3
# @Date    : 2021-10-27
# @Author  : Serafim Nenarokov (serafim.nenarokov@gmail.com)

from argparse import ArgumentParser
from pprint import pprint

import os
import openpyxl
from pathlib import Path
from jinja2 import Template

from lib.record import Record


def read_args():
  parser = ArgumentParser(description = 'Gets GC statistics from fasta file and annotation GFF results')
  parser.add_argument('-i', '--input', required = True, help = 'path to XLSX file')
  parser.add_argument('-o', '--output', required = True, help = 'path to output folder for XML files')
  return parser.parse_args()

def check_output_folder(val):
  if not os.path.isdir(val):
    print("Output path is not a directory or doesn't exist.")
    exit(-1)

def get_entries(sheet):
  result = []

  col_names = []
  for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)

  for row in sheet.iter_rows(min_row=2):
    obj = {}

    for i, cell in enumerate(row):
      val = cell.value

      if val:
        obj[col_names[i]] = val

    if len(obj) > 0:
      result.append(obj)

  return result


def main():
  args = read_args()
  check_output_folder(args.output)

  xlsx_path = Path(args.input)
  obj = openpyxl.load_workbook(xlsx_path)

  data_sheet = obj.active

  records = []

  for entry in get_entries(data_sheet):
    record = Record(entry)
    records.append(record)

    record.validate()
    if not record.is_valid():
      record.print_errors()

  ## working with template
  with open('template.xml') as f:
    template = Template(f.read())

  for record in records:
    text = template.render(record=record)
    out_path = os.path.join(args.output, f'{record.invoice_number}.xml')
    
    with open(out_path, 'w') as out_f:
      out_f.write(text)


if __name__ == '__main__':
  main()